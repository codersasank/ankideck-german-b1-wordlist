import openpyxl, requests, json, base64

url = "https://texttospeech.googleapis.com/v1beta1/text:synthesize"

def synthesize_text(filename, txt, voice_name):
    data = {"input": {"text": txt},
            "voice": {"name":  voice_name, "languageCode": "de-DE"},
            "audioConfig": {"audioEncoding": "MP3"}};
    headers = {"content-type": "application/json", "X-Goog-Api-Key": "your-key-here" }
    response = requests.post(url=url, json=data, headers=headers)
    if response.status_code == 200:
        # Save the audio content as an MP3 file
        with open('audio/'+filename+".mp3", "wb") as out:
            content = json.loads(response.content)
            audio_content_base64 = content["audioContent"]
            audio_content_bytes = base64.b64decode(audio_content_base64)
            out.write(audio_content_bytes)
    else:
        print(f"Request failed with status code: {response.status_code}")
        print(response.text)
            
if __name__=="__main__":
    # Give the location of the XLSX file
    path = "A2_Wortliste.xlsx"
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    
    for row_idx in range(2,1882):
        de_word = sheet_obj.cell(row=row_idx, column=2)
        de_article = sheet_obj.cell(row=row_idx, column=3)
        de_word_with_article = None
        if (de_article.value is None) or (de_article.value.strip()==""):
            de_word_with_article = de_word.value.strip()
        else:
            de_word_with_article = de_article.value.strip() + ' ' + de_word.value.strip()      
        de_sentence = sheet_obj.cell(row=row_idx, column=7)
        note_id = sheet_obj.cell(row=row_idx, column=1)
        synthesize_text('word_' + note_id.value, de_word_with_article, "de-DE-Wavenet-B" )
        synthesize_text('sentence_' + note_id.value, de_sentence.value, "de-DE-Neural2-F")
