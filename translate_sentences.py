import openpyxl, deepl

auth_key = "your-key-here"
translator = deepl.Translator(auth_key)

def translate(text):
    result = translator.translate_text(text, source_lang="DE", target_lang="EN-GB")
    print (result)
    return result.text
    
if __name__=="__main__":
    # Give the location of the XLSX file
    path = "B1_Wortliste.xlsx"
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    arial = openpyxl.styles.Font(name='Arial', size=12)
    left_align= openpyxl.styles.alignment.Alignment(horizontal='left', vertical="bottom")
    
    for row_idx in range(2,1882):
        de_sentence = sheet_obj.cell(row=row_idx, column=7)
        en_sentence_obj = sheet_obj.cell(row=row_idx, column=9)

        en_sentence = translate(de_sentence.value)
        en_sentence_obj.value = en_sentence
        en_sentence_obj.font = arial
        en_sentence_obj.alignment = left_align
    # Instead of overwriting the same file, generate a new XLSX file with the translations
    wb_obj.save("test.xlsx")
